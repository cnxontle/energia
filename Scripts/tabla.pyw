import pymsgbox
import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

def filtrar_filas(columna,filtro):
    global df
    columna_filtro = None
    for column in df.columns:
        if columna in column.upper():
            columna_filtro = column
            break
    if columna_filtro is None:
        print(f"La columna '{columna}' no se encontró en el archivo Excel.")
    else:
        if filtro == "N/A":
            df = df.dropna(subset=[columna_filtro])
        else:
            df = df[df[columna_filtro] == filtro]

# Cargar archivo
script_dir = os.path.dirname(os.path.abspath(__file__))
destino = os.path.abspath(os.path.join(script_dir, os.pardir))
RutaMadre = destino + "\\"
os.chdir(RutaMadre)

excel_file = RutaMadre + "CUOTA ENERGETICA.xlsm"


#  LEER CELDAS CON NOMBRE
wb = load_workbook(excel_file, data_only=True)

def leer_nombre(nombre):
    ref = wb.defined_names[nombre].attr_text
    hoja, rango = ref.split('!')
    hoja = wb[hoja.strip("'")]
    return hoja[rango].value

valor_estado       = leer_nombre("ESTADO")
valor_oref         = leer_nombre("ESTADO")
valor_clave_estado = leer_nombre("CLAVE_ESTADO")
valor_cader        = leer_nombre("CADER")
valor_clave_cader  = leer_nombre("CLAVE_CADER")
valor_clave_ddr    = leer_nombre("CLAVE_DISTRITO")
valor_ddr          = leer_nombre("DISTRITO")
valor_vacio       = ""

# --- Cargar DataFrame ---
df = pd.read_excel(excel_file)

# ===== FILTROS =====
filtrar_filas('ESTATUS','N/A')
filtrar_filas('ESTADO DEL PERMISO','VIGENTE')

columnas_objetivo = ['CLAVE MUNICIPIO','MUNICIPIO', 'CLAVE LOCALIDAD','LOCALIDAD', 'CLAVE DE REGISTRO (PEUA)', 'TIPO DE PERSONA', 'NOMBRE (S)', 'APELLIDO PATERNO', 'APELLIDO MATERNO', 'CURP', 'RFC', 
                      'HOMBRES','MUJERES','PRODUCTORES BENEFICIADOS', 'RPU', 'NO. DE CUENTA', 'RMU', 'TIPO DE DOCUMENTO QUE ACREDITA EL USO Y APROVECHAMIENTO DE AGUA',
                      'NUMERO DE FOLIO DEL DOCUMENTO PRESENTADO', 'VIGENCIA DEL TITULO DE CONCESION  (dd/mm/aaaa)', 'VOLUMEN CONCECIONADO',
                      'LATITUD', 'LONGITUD', 'HP DEL EQUIPO REGISTRADO', 'CUOTA ENERGETICA CALCULADA', 'SUP BENEFICIADA',
                      'CULTIVO', 'RENDIMIENTO (TON/HA)', 'FECHA DE ACTUALIZACION', 'TELEFONO', 'CORREO ELECTRONICO']

df = df[columnas_objetivo]

# FORMATO
df['RPU'] = df['RPU'].astype(str).str.replace('.0', '')
df['NUMERO DE FOLIO DEL DOCUMENTO PRESENTADO'] = df['NUMERO DE FOLIO DEL DOCUMENTO PRESENTADO'].astype(str)
df['FECHA DE SOLICITUD'] = pd.to_datetime(df['FECHA DE ACTUALIZACION'], errors='coerce').dt.strftime('%d/%m/%Y')
df = df.drop(columns=['FECHA DE ACTUALIZACION'])

# CONVERTIR A STR 
df['MUNICIPIO'] = df['MUNICIPIO'].astype(str)
df['LOCALIDAD'] = df['LOCALIDAD'].astype(str)
df['NOMBRE (S)'] = df['NOMBRE (S)'].astype(str)
df['APELLIDO PATERNO'] = df['APELLIDO PATERNO'].astype(str)
df['APELLIDO MATERNO'] = df['APELLIDO MATERNO'].astype(str)
df['CURP'] = df['CURP'].astype(str)
df['RFC'] = df['RFC'].astype(str)
df['TIPO DE PERSONA'] = df['TIPO DE PERSONA'].astype(str)
df['NO. DE CUENTA'] = df['NO. DE CUENTA'].astype(str)
df['RMU'] = df['RMU'].astype(str)
df['CULTIVO'] = df['CULTIVO'].astype(str)
df['CLAVE DE REGISTRO (PEUA)'] = df['CLAVE DE REGISTRO (PEUA)'].astype(str)

# CONVERTIR A MAYUSCULAS
df['MUNICIPIO'] = df['MUNICIPIO'].str.upper()
df['LOCALIDAD'] = df['LOCALIDAD'].str.upper()
df['NOMBRE (S)'] = df['NOMBRE (S)'].str.upper()
df['APELLIDO PATERNO'] = df['APELLIDO PATERNO'].str.upper()
df['APELLIDO MATERNO'] = df['APELLIDO MATERNO'].str.upper()
df['CURP'] = df['CURP'].str.upper()
df['RFC'] = df['RFC'].str.upper()
df['TIPO DE PERSONA'] = df['TIPO DE PERSONA'].str.upper()
df['TIPO DE PERSONA'] = df['TIPO DE PERSONA'].replace({'F': 'FISICA', 'M': 'MORAL'})
df['NO. DE CUENTA'] = df['NO. DE CUENTA'].str.upper()
df['RMU'] = df['RMU'].str.upper()
df['CULTIVO'] = df['CULTIVO'].str.upper()
df['CLAVE DE REGISTRO (PEUA)'] = df['CLAVE DE REGISTRO (PEUA)'].str.upper()


df['TIPO DE PERSONA'] = df['TIPO DE PERSONA'].replace({'F':'FISICA','M':'MORAL'})


#   AGREGAR VALORES LEÍDOS
df.insert(0, "NUM", range(1, len(df) + 1))
df.insert(1, "OREF", valor_oref)
df.insert(2, "CVE ESTADO", valor_clave_estado)
df.insert(3, "ESTADO", valor_estado)
df.insert(4, "CLAVE_DDR", valor_clave_ddr)
df.insert(5, "DDR", valor_ddr)
df.insert(6, "CLAVE_CADER", valor_clave_cader)
df.insert(7, "CADER", valor_cader)
df.insert(19, "REGIMEN SAT", valor_vacio)
df.insert(26, "FOLIO TITULO", valor_vacio)
df.insert(27, "VIGENCIA TITULO", valor_vacio)
df.insert(36, "GASTO DEL EQUIPO DE BOMBEO", valor_vacio)
df.insert(41, "TELEFONO OFICINA", valor_vacio)

#mover la columna fecha de solicitud a la posicion 40
fecha_solicitud = df.pop('FECHA DE SOLICITUD')
df.insert(40, 'FECHA DE SOLICITUD', fecha_solicitud)

import pymsgbox
import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

def filtrar_filas(columna,filtro):
    global df
    columna_filtro = None
    for column in df.columns:
        if columna in column.upper():
            columna_filtro = column
            break
    if columna_filtro is None:
        print(f"La columna '{columna}' no se encontró en el archivo Excel.")
    else:
        if filtro == "N/A":
            df = df.dropna(subset=[columna_filtro])
        else:
            df = df[df[columna_filtro] == filtro]

# Cargar archivo
script_dir = os.path.dirname(os.path.abspath(__file__))
destino = os.path.abspath(os.path.join(script_dir, os.pardir))
RutaMadre = destino + "\\"
os.chdir(RutaMadre)

excel_file = RutaMadre + "CUOTA ENERGETICA.xlsm"


#  LEER CELDAS CON NOMBRE
wb = load_workbook(excel_file, data_only=True)

def leer_nombre(nombre):
    ref = wb.defined_names[nombre].attr_text
    hoja, rango = ref.split('!')
    hoja = wb[hoja.strip("'")]
    return hoja[rango].value

valor_estado       = leer_nombre("ESTADO")
valor_oref         = leer_nombre("ESTADO")
valor_clave_estado = leer_nombre("CLAVE_ESTADO")
valor_cader        = leer_nombre("CADER")
valor_clave_cader  = leer_nombre("CLAVE_CADER")
valor_clave_ddr    = leer_nombre("CLAVE_DISTRITO")
valor_ddr          = leer_nombre("DISTRITO")
valor_vacio       = ""

# --- Cargar DataFrame ---
df = pd.read_excel(excel_file)

# ===== FILTROS =====
filtrar_filas('ESTATUS','N/A')
filtrar_filas('ESTADO DEL PERMISO','VIGENTE')

columnas_objetivo = ['CLAVE MUNICIPIO','MUNICIPIO', 'CLAVE LOCALIDAD','LOCALIDAD', 'CLAVE DE REGISTRO (PEUA)', 'TIPO DE PERSONA', 'NOMBRE (S)', 'APELLIDO PATERNO', 'APELLIDO MATERNO', 'CURP', 'RFC', 
                      'HOMBRES','MUJERES','PRODUCTORES BENEFICIADOS', 'RPU', 'NO. DE CUENTA', 'RMU', 'TIPO DE DOCUMENTO QUE ACREDITA EL USO Y APROVECHAMIENTO DE AGUA',
                      'NUMERO DE FOLIO DEL DOCUMENTO PRESENTADO', 'VIGENCIA DEL TITULO DE CONCESION  (dd/mm/aaaa)', 'VOLUMEN CONCECIONADO',
                      'LATITUD', 'LONGITUD', 'HP DEL EQUIPO REGISTRADO', 'CUOTA ENERGETICA CALCULADA', 'SUP BENEFICIADA',
                      'CULTIVO', 'RENDIMIENTO (TON/HA)', 'FECHA DE ACTUALIZACION', 'TELEFONO', 'CORREO ELECTRONICO']

df = df[columnas_objetivo]

# FORMATO
df['RPU'] = df['RPU'].astype(str).str.replace('.0', '')
df['NUMERO DE FOLIO DEL DOCUMENTO PRESENTADO'] = df['NUMERO DE FOLIO DEL DOCUMENTO PRESENTADO'].astype(str)
df['FECHA DE SOLICITUD'] = pd.to_datetime(df['FECHA DE ACTUALIZACION'], errors='coerce').dt.strftime('%d/%m/%Y')
df = df.drop(columns=['FECHA DE ACTUALIZACION'])

# CONVERTIR A STR 
df['MUNICIPIO'] = df['MUNICIPIO'].astype(str)
df['LOCALIDAD'] = df['LOCALIDAD'].astype(str)
df['NOMBRE (S)'] = df['NOMBRE (S)'].astype(str)
df['APELLIDO PATERNO'] = df['APELLIDO PATERNO'].astype(str)
df['APELLIDO MATERNO'] = df['APELLIDO MATERNO'].astype(str)
df['CURP'] = df['CURP'].astype(str)
df['RFC'] = df['RFC'].astype(str)
df['TIPO DE PERSONA'] = df['TIPO DE PERSONA'].astype(str)
df['NO. DE CUENTA'] = df['NO. DE CUENTA'].astype(str)
df['RMU'] = df['RMU'].astype(str)
df['CULTIVO'] = df['CULTIVO'].astype(str)
df['CLAVE DE REGISTRO (PEUA)'] = df['CLAVE DE REGISTRO (PEUA)'].astype(str)

# CONVERTIR A MAYUSCULAS
df['MUNICIPIO'] = df['MUNICIPIO'].str.upper()
df['LOCALIDAD'] = df['LOCALIDAD'].str.upper()
df['NOMBRE (S)'] = df['NOMBRE (S)'].str.upper()
df['APELLIDO PATERNO'] = df['APELLIDO PATERNO'].str.upper()
df['APELLIDO MATERNO'] = df['APELLIDO MATERNO'].str.upper()
df['CURP'] = df['CURP'].str.upper()
df['RFC'] = df['RFC'].str.upper()
df['TIPO DE PERSONA'] = df['TIPO DE PERSONA'].str.upper()
df['TIPO DE PERSONA'] = df['TIPO DE PERSONA'].replace({'F': 'FISICA', 'M': 'MORAL'})
df['NO. DE CUENTA'] = df['NO. DE CUENTA'].str.upper()
df['RMU'] = df['RMU'].str.upper()
df['CULTIVO'] = df['CULTIVO'].str.upper()
df['CLAVE DE REGISTRO (PEUA)'] = df['CLAVE DE REGISTRO (PEUA)'].str.upper()


df['TIPO DE PERSONA'] = df['TIPO DE PERSONA'].replace({'F':'FISICA','M':'MORAL'})


#   AGREGAR VALORES LEÍDOS
df.insert(0, "NUM", range(1, len(df) + 1))
df.insert(1, "OREF", valor_oref)
df.insert(2, "CVE ESTADO", valor_clave_estado)
df.insert(3, "ESTADO", valor_estado)
df.insert(4, "CLAVE_DDR", valor_clave_ddr)
df.insert(5, "DDR", valor_ddr)
df.insert(6, "CLAVE_CADER", valor_clave_cader)
df.insert(7, "CADER", valor_cader)
df.insert(19, "REGIMEN SAT", valor_vacio)
df.insert(26, "FOLIO TITULO", valor_vacio)
df.insert(27, "VIGENCIA TITULO", valor_vacio)
df.insert(36, "GASTO DEL EQUIPO DE BOMBEO", valor_vacio)
df.insert(41, "TELEFONO OFICINA", valor_vacio)

#mover la columna fecha de solicitud a la posicion 40
fecha_solicitud = df.pop('FECHA DE SOLICITUD')
df.insert(40, 'FECHA DE SOLICITUD', fecha_solicitud)

# Filtro: solo filas donde el tipo de documento es "TITULO DE CONCESIÓN"
mask_titulo = df['TIPO DE DOCUMENTO QUE ACREDITA EL USO Y APROVECHAMIENTO DE AGUA'] == "TITULO DE CONCESIÓN"

# Llenar "FOLIO TITULO"
df.loc[mask_titulo, "FOLIO TITULO"] = df.loc[mask_titulo, "NUMERO DE FOLIO DEL DOCUMENTO PRESENTADO"]

# Llenar "VIGENCIA TITULO"
df.loc[mask_titulo, "VIGENCIA TITULO"] = df.loc[mask_titulo, "VIGENCIA DEL TITULO DE CONCESION  (dd/mm/aaaa)"]

# eliminar el contenido de celdas "NUMERO DE FOLIO DEL DOCUMENTO PRESENTADO" y "VIGENCIA DEL TITULO DE CONCESION  (dd/mm/aaaa)" en donde se especifique "TITULO DE CONCESIÓN"
df.loc[mask_titulo, "NUMERO DE FOLIO DEL DOCUMENTO PRESENTADO"] = ""
df.loc[mask_titulo, "VIGENCIA DEL TITULO DE CONCESION  (dd/mm/aaaa)"] = ""
df.loc[mask_titulo, "TIPO DE DOCUMENTO QUE ACREDITA EL USO Y APROVECHAMIENTO DE AGUA"] = ""

#en la columna "TIPO DE DOCUMENTO QUE ACREDITA EL USO Y APROVECHAMIENTO DE AGUA", reemplazar los valores "LIBRE ALUMBRAMIENTO" por "CONSTANCIA DE REGISTRO DE OBRA DE ALUMBRAMIENTO"
df['TIPO DE DOCUMENTO QUE ACREDITA EL USO Y APROVECHAMIENTO DE AGUA'] = df['TIPO DE DOCUMENTO QUE ACREDITA EL USO Y APROVECHAMIENTO DE AGUA'].replace({'LIBRE ALUMBRAMIENTO': 'CONSTANCIA DE REGISTRO DE OBRA DE ALUMBRAMIENTO'})

# Guardar archivo
fecha_actual = datetime.now().strftime("%d_%m_%Y")
nombre_archivo = f"padron_{fecha_actual}.xlsx"

df.to_excel(nombre_archivo, index=False)
pymsgbox.alert('Procedimiento completado!', 'EXPORTAR TABLA')


# Guardar archivo
fecha_actual = datetime.now().strftime("%d_%m_%Y")
nombre_archivo = f"padron_{fecha_actual}.xlsx"

df.to_excel(nombre_archivo, index=False)
pymsgbox.alert('Procedimiento completado!', 'EXPORTAR TABLA')

