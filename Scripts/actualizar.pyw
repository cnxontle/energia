import os
import zipfile
import requests
import openpyxl
import pandas as pd
import xlwings as xw
from time import sleep
import pymsgbox

sleep (3)
# Actualizar repositorio local omitiendo archivos de configuracion
def descargar_y_descomprimir_zip(url_zip, destino, rec):
    response = requests.get(url_zip)
    if response.status_code == 200:
        with open("programa_actualizado.zip", "wb") as zip_file:
            zip_file.write(response.content)
        with zipfile.ZipFile("programa_actualizado.zip", "r") as zip_ref:
            for file_name in zip_ref.namelist():
                if not file_name.endswith('.txt') and not file_name.endswith('.md') and not file_name.endswith('.gitignore'):
                    zip_ref.extract(file_name, destino)
        os.remove("programa_actualizado.zip")
    else:
        pymsgbox.alert("Error al descargar y descomprimir el repositorio", title='Error')
        os.remove(rec)
        exit()

# Hacer respaldo de la informacion del excel
def actualizar_excel(filename,nombre,min):
    wb = openpyxl.load_workbook(filename, data_only=False, keep_vba=True)
    sheet = wb[nombre]
    max_row = sheet.max_row
    contenido_columnas = []
    for row in sheet.iter_rows(min_row=min, max_row=max_row, min_col=min, max_col=64):  # Columnas B a BH
        fila = [cell.value for cell in row]
        contenido_columnas.append(fila)
    df = pd.DataFrame(contenido_columnas)
    return df

# Obtener el directorio del script
script_dir = os.path.dirname(os.path.abspath(__file__))
destino = os.path.abspath(os.path.join(script_dir, os.pardir))
destino2 = os.path.abspath(os.path.join(destino, os.pardir))
archivo_xlsm = os.path.join(destino, 'CUOTA ENERGETICA.xlsm')

# Obtener datos de excel
contenido_del_dataframe = actualizar_excel(archivo_xlsm,"DATOS",2)

# Crear un archivo Excel "recuperado.xlsx"
archivo_excel_recuperado = os.path.join(script_dir, 'recuperado.xlsx')
rec = openpyxl.Workbook()

# Crear hojas
ws_padron_cfe = rec.create_sheet(title="padron cfe")
ws_padron_cfe = rec.create_sheet(title="REPORTE")
ws_datos = rec.create_sheet(title="DATOS")

# Escribir respaldo
for row in contenido_del_dataframe.values.tolist():
    ws_datos.append(row)

# Guardar el libro de trabajo
rec.save(archivo_excel_recuperado)
rec.close()

# Reemplazar repositorio
url_zip = "https://github.com/cnxontle/energia/archive/main.zip"
descargar_y_descomprimir_zip(url_zip, destino2, archivo_excel_recuperado)

# Modificar el nuevo excel
app = xw.App(visible=False)
wb = app.books.open(archivo_xlsm)

# Restaurar datos
sheet = wb.sheets['DATOS']
sheet.range('B2').value = contenido_del_dataframe.values

# Guarda los cambios en el archivo
wb.save()
wb.close()
app.quit()

# eliminar el archivo recuperado
os.remove(archivo_excel_recuperado)

pymsgbox.alert("Procedimiento Completado!",title='Actualicación del programa')