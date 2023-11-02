import os
import zipfile
import requests
import openpyxl
import pandas as pd
import xlwings as xw
import psutil
import win32com.client

def is_excel_open(file_path):
    file_name = os.path.basename(file_path).lower()
    for process in psutil.process_iter(attrs=['pid', 'name']):
        try:
            if 'excel' in process.info['name'].lower():
                for window in process.open_files():
                    if file_name in window.path.lower():
                        return True
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            pass
    return False

def save_and_close_excel(file_path):
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        workbook = excel.Workbooks.Open(file_path)
        workbook.Save()
        workbook.Close()
        excel.Quit()
    except Exception as e:
        print(f"Error al guardar y cerrar Excel: {str(e)}")

def descargar_y_descomprimir_zip(url_zip, destino):
    response = requests.get(url_zip)
    if response.status_code == 200:
        with open("programa_actualizado.zip", "wb") as zip_file:
            zip_file.write(response.content)
        with zipfile.ZipFile("programa_actualizado.zip", "r") as zip_ref:
            zip_ref.extractall(destino)
        os.remove("programa_actualizado.zip")
        print("Programa actualizado correctamente.")
    else:
        print("Error al descargar el archivo ZIP")

def actualizar_excel(filename,nombre,min):
    wb = openpyxl.load_workbook(filename, data_only=False, keep_vba=True)
    sheet = wb[nombre]
    max_row = sheet.max_row
    contenido_columnas = []
    for row in sheet.iter_rows(min_row=min, max_row=max_row, min_col=min, max_col=54):  # Columnas B a BB
        fila = [cell.value for cell in row]
        contenido_columnas.append(fila)
    df = pd.DataFrame(contenido_columnas)
    return df

# Obtener el directorio del script
script_dir = os.path.dirname(os.path.abspath(__file__))
archivo_xlsm = os.path.join(script_dir, 'CUOTA ENERGETICA.xlsm')

if is_excel_open(archivo_xlsm):
    save_and_close_excel(archivo_xlsm)

# # obtener el contenido de excel
contenido_del_dataframe = actualizar_excel(archivo_xlsm,"DATOS",2)
contenido_del_dataframe2 = actualizar_excel(archivo_xlsm,"bombas",1)

# Reemplazar repositorio
url_zip = "https://github.com/cnxontle/energia/archive/main.zip"
script_dir = os.path.dirname(os.path.abspath(__file__))
destino = os.path.abspath(os.path.join(script_dir, os.pardir))
descargar_y_descomprimir_zip(url_zip, destino)

# Accede modificar el excel
wb = xw.Book(archivo_xlsm)

# restaurar datos
sheet = wb.sheets['DATOS']
sheet.range('B2').value = contenido_del_dataframe.values

#restaurar bombas
sheet2 = wb.sheets['bombas']
sheet2.range('a1').value = contenido_del_dataframe2.values

# Guarda los cambios en el archivo
wb.save()
wb.close()